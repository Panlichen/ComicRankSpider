# -*- coding: UTF-8 -*-
'''
Created on 2017年8月13日

@author: mlcp
'''
from openpyxl import Workbook
import openpyxl
from selenium import webdriver

class U17:
    
    SAVE_PATH = None
    
    def __init__(self, path):
        self.SAVE_PATH = path
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        ws['J2'] = '排名'
        ws['K2'] = '新签约榜'
        ws['L2'] = '类型'
        ws['M2'] = '签约读者数（总点击）'
        ws['N2'] = '圣殿榜'
        ws['O2'] = '类型'
        ws['P2'] = '总点击'
        ws['Q2'] = '总月票'
        ws['R2'] = '订阅上升榜'
        ws['S2'] = '类型'
        wb.save(self.SAVE_PATH)
        
    def crawl(self):
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        driver = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        driver2 = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        
        
        '''
        新签约榜
        '''
        
        driver.get('http://comic.u17.com/rank/t1.html')
        names = driver.find_elements_by_css_selector('div.img_box .categray')
        idx = 1
        for name in names:
            a_ele = name.find_element_by_css_selector('a')
            d_ele = name.find_element_by_css_selector('div')
            ws['J' + str(idx + 2)] = idx
            ws['K' + str(idx + 2)] = a_ele.text
            ws['L' + str(idx + 2)] = d_ele.text
            driver2.get(a_ele.get_attribute('href'))
            ws['M' + str(idx + 2)] = driver2.find_element_by_css_selector('body > div.wrap.cf > div.comic_info > div.left > div.info > div.top > div.line1 > i').text
            
            print 'U17 新签约榜 '+ idx + '/20'
            if idx == 20:
                break
            idx += 1
        wb.save(self.SAVE_PATH)
            
        '''
        圣殿榜
        '''
        path = driver.find_element_by_css_selector('body > div.wrap.cf > div.ranking_left > div:nth-child(2) > a:nth-child(2)').get_attribute('href')
        driver.get(path)
        names = driver.find_elements_by_css_selector('div.img_box .categray')
        idx = 1
        for name in names:
            a_ele = name.find_element_by_css_selector('a')
            d_ele = name.find_element_by_css_selector('div')
            ws['N' + str(idx + 2)] = a_ele.text
            ws['O' + str(idx + 2)] = d_ele.text
            driver2.get(a_ele.get_attribute('href'))
            if idx == 1:
                ws['P' + str(idx + 2)] = driver2.find_element_by_css_selector('#comicInfo > div.zhenhunjie_wrap.bg1 > div.area > div.wrap.cf > div.comic_info > div.left > div.info > div.top > div.cf.line2 > div:nth-child(2) > span').text
                ws['Q' + str(idx + 2)] = driver2.find_element_by_css_selector('#comicInfo > div.zhenhunjie_wrap.bg1 > div.area > div.wrap.cf > div.comic_info > div.left > div.info > div.top > div.cf.line2 > div:nth-child(3) > span').text
            else:
                ws['P' + str(idx + 2)] = driver2.find_element_by_css_selector('body > div.wrap.cf > div.comic_info > div.left > div.info > div.top > div.line1 > i').text
                ws['Q' + str(idx + 2)] = driver2.find_element_by_css_selector('body > div.wrap.cf > div.comic_info > div.left > div.info > div.top > div.line2 > i').text
            
            print 'U17 圣殿榜 '+ idx + '/20'
            if idx == 20:
                break
            idx += 1
        wb.save(self.SAVE_PATH)
        
        '''
        订阅上升榜
        '''
        
        path = driver.find_element_by_css_selector('body > div.wrap.cf > div.ranking_left > div:nth-child(2) > a:nth-child(3)').get_attribute('href')
        driver.get(path)
        names = driver.find_elements_by_css_selector('div.img_box .categray')
        idx = 1
        for name in names:
            a_ele = name.find_element_by_css_selector('a')
            d_ele = name.find_element_by_css_selector('div')
            ws['R' + str(idx + 2)] = a_ele.text
            ws['S' + str(idx + 2)] = d_ele.text
            
            print 'U17 订阅上升榜 '+ idx + '/20'
            if idx == 20:
                break
            idx += 1
        wb.save(self.SAVE_PATH)
        
        driver.close()
        driver2.close()
        