# -*- coding: UTF-8 -*-
'''
Created on 2017年8月13日

@author: mlcp
'''
from openpyxl import Workbook
import openpyxl
from selenium import webdriver

class NetEase:
    
    SAVE_PATH = None

    def __init__(self, path):
        self.SAVE_PATH = path
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        ws['T2'] = '排名'
        ws['U2'] = '人气榜'
        ws['V2'] = '类型'
        ws['W2'] = '周点击'
        ws['X2'] = '打赏榜'
        ws['Y2'] = '类型'
        ws['Z2'] = '粉丝值'
        wb.save(self.SAVE_PATH)
        
    def crawl(self):
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        driver = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        driver2 = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        
        '''
        人气榜
        '''
        
        path = 'https://manhua.163.com/rank/list.do?type=-1&periodType=1'
        driver.get(path)
        names = driver.find_elements_by_css_selector('.sub-rank-wrap.f-cb .comic-desc')
        idx = 1
        for name in names:
            ws['T' + str(idx + 2)] = idx
            str_name = name.find_element_by_css_selector('.comic-title').text
            ws['U' + str(idx + 2)] = str_name
            
            path = name.find_element_by_css_selector('a:nth-child(1)').get_attribute('href')
            str_tags = self.get_str_tag(path, driver2)
            ws['V' + str(idx + 2)] = str_tags
            
            str_click = name.find_element_by_css_selector('.comic-meta').text
            ws['W' + str(idx + 2)] = str_click[str_click.find(' ') + 1 : ]
            
            print 'NetEase 人气榜 ', idx , '/20'
            if idx == 20:
                break
            idx += 1
            
        wb.save(self.SAVE_PATH)
        
        '''
        打赏榜
        '''
        path = 'https://manhua.163.com/rank/list.do?type=-2&periodType=1'
        driver.get(path)
        names = driver.find_elements_by_css_selector('.sub-rank-wrap.f-cb .comic-desc')
        idx = 1
        for name in names:
            str_name = name.find_element_by_css_selector('.comic-title').text
            ws['X' + str(idx + 2)] = str_name
            
            path = name.find_element_by_css_selector('a:nth-child(1)').get_attribute('href')
            str_tags = self.get_str_tag(path, driver2)
            ws['Y' + str(idx + 2)] = str_tags
            
            str_award = name.find_element_by_css_selector('.comic-meta').text
            ws['Z' + str(idx + 2)] = str_award[4 : ]
            print 'NetEase 打赏榜 ', idx , '/20'
            if idx == 20:
                break
            idx += 1
            
        wb.save(self.SAVE_PATH)
        
        driver.close()
        driver2.close()
            
    def get_str_tag(self, path, driver):
        driver.get(path)
        tags = driver.find_elements_by_css_selector('body > div.sr-view > div.sr-main > div.sr-info > div.sr-detail.f-pr > div.sr-detail__middle.js-detail-middle > dl > dd:nth-child(4) a')
        str_tags = ""
        for tag in tags:
            str_tags += tag.text + ' '
        return str_tags
        
            