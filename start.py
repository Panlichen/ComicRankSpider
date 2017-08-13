# -*- coding: UTF-8 -*-
'''
Created on 2017年8月13日

@author: mlcp
'''
from Tencent import Tencent
from U17 import U17
from NetEase import NetEase
import time
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
import openpyxl

if __name__ == '__main__':
    path = '/Users/mlcp/PKU/17暑假/spider/' + 'ComicsRank-' + time.strftime("%Y-%m-%d", time.localtime()) + '.xlsx'
    wb = Workbook()
    ws = wb.active
    
    ws['A1'] = '腾讯'
    ws['J1'] = '有妖气'
    ws['T1'] = '网易漫画'
    ws.merge_cells('A1:I1')
    ws.merge_cells('J1:S1')
    ws.merge_cells('T1:Z1')
    
    myfont = Font(b=True)
    al = Alignment(horizontal="center", vertical="center")
    ws['A1'].font = myfont
    ws['A1'].alignment = al
    ws['J1'].font = myfont
    ws['J1'].alignment = al
    ws['T1'].font = myfont
    ws['T1'].alignment = al
    wb.save(path)
    
    tencent = Tencent(path)
    tencent.crawl()
    print "腾讯完成"
    
    u17 = U17(path)
    u17.crawl()
    print "有妖气完成"
    '''
    netease = NetEase(path) 
    netease.crawl()
    print "网易完成"  
    '''