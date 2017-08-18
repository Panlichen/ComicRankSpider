# -*- coding: UTF-8 -*-
'''
Created on 2017年8月13日

@author: mlcp
'''
from openpyxl import Workbook
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.action_chains import ActionChains
from numpy import str_

class Tencent(object):
    SAVE_PATH = None

    def __init__(self, path):
        self.SAVE_PATH = path 
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        ws['A2'] = "排名"
        ws['B2'] = "月票榜"
        ws['C2'] = "类型"
        ws['D2'] = "本月月票数"
        ws['E2'] = "打赏榜"
        ws['F2'] = "类型"
        ws['G2'] = "人气榜"
        ws['H2'] = "类型"
        ws['I2'] = "总人气"
        wb.save(self.SAVE_PATH)
    
    def crawl(self):
        wb = openpyxl.load_workbook(self.SAVE_PATH)
        ws = wb.active
        driver = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        driver2 = webdriver.PhantomJS(executable_path='/usr/local/bin/phantomjs')
        
        '''
        月票
        '''
        
        driver.get('http://ac.qq.com/Rank')
        names = driver.find_elements_by_css_selector('#month_ticket_rank_content a')
        rank_nums = driver.find_elements_by_css_selector('#month_ticket_rank_content span')
        idx = 1
        for name in names:
            ws['A' + str(idx + 2)] = idx
            ws['B' + str(idx + 2)] = name.get_attribute('title')
            
            str_tag = self.get_str_tag(name.get_attribute('href'), driver2)
            ws['C' + str(idx + 2)] = str_tag
            
            print 'Tencent 月票 ', idx , '/20'
            
            if idx == 20:
                break
            idx += 1
        idx = 1
        for rank_num in rank_nums:
            ws['D' + str(idx + 2)] = int(rank_num.text)
            if idx == 20:
                break
            idx += 1
        path = driver.find_element_by_css_selector('body > div.mod-rank-wrap.clearfix > div.mod-rank-menu.ui-left > div.mod-rank-tab-wrap > div > ul:nth-child(2) > li:nth-child(3) > a').get_attribute('href')
        wb.save(self.SAVE_PATH)
        
        '''
        打赏
        '''
        
        driver.get(path)
        names = driver.find_elements_by_css_selector('div.rank-tabcon a')
        idx = 1
        for name in names:
            ws['E' + str(idx + 2)] = name.get_attribute('title')
            print 'Tencent 打赏 ', idx , '/20'
            str_tag = self.get_str_tag(name.get_attribute('href'), driver2)
            ws['F' + str(idx + 2)] = str_tag
            if idx == 20:
                break
            idx += 1    
            
        path = driver.find_element_by_css_selector('body > div.mod-rank-wrap.clearfix.rank-sub-mb > div.mod-rank-menu.ui-left > div.mod-rank-tab-wrap > div > ul:nth-child(2) > li:nth-child(4) > a').get_attribute('href')
        wb.save(self.SAVE_PATH)
        '''
        人气
        '''
        
        driver.get(path)
        week_btn = driver.find_element_by_css_selector('#month_all_rank > a:nth-child(2)')
        #print week_btn.text, week_btn.size
        ActionChains(driver).move_to_element(week_btn).perform()
        driver.execute_script('$(arguments[0]).click();',week_btn)
        driver.execute_script('var evObj = document.createEvent("MouseEvents");evObj.initMouseEvent("mouseover",true, false, window, 0, 0, 0, 0, 0, false, false, false, false, 0, null);arguments[0].dispatchEvent(evObj);',week_btn)
        driver.execute_script("$('#month_all_rank > a:nth-child(2)').mouseover()", week_btn)
        idx = 1
        names = driver.find_elements_by_css_selector('div.rank-tabcon a')
        for name in names:
            ws['G' + str(idx + 2)] = name.get_attribute('title')
            
            path = name.get_attribute('href')
            driver2.get(path)
            tags = driver2.find_elements_by_css_selector('#tags-show a')
            str_tag = ""
            for tag in tags:
                str_tag += tag.get_attribute('title') + ' '
            ws['H' + str(idx + 2)] = str_tag
            data = driver2.find_element_by_css_selector('#special_bg > div:nth-child(3) > div.ui-left.works-intro-wr > div.works-intro.clearfix > div.works-intro-detail.ui-left > div.works-intro-text > p.works-intro-digi > span:nth-child(2) > em').text
            ws['I' + str(idx + 2)] = data
            print 'Tencent 人气 ', idx , '/20'
            if idx == 20:
                break
            idx += 1
            
        wb.save(self.SAVE_PATH)
        
        driver.close()
        driver2.close()
             
    def get_str_tag(self, path, driver):
        driver.get(path)
        tags = driver.find_elements_by_css_selector('#tags-show a')
        str_tag = ""
        for tag in tags:
            str_tag += tag.get_attribute('title') + ' '
        return str_tag
        
            